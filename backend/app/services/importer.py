"""交割单解析与配对引擎

支持文件：
  - Excel (.xlsx)
  - CSV / 制表符分隔文本 (.csv / .txt)

兼容性策略（不针对某家软件写死）：
  1. 自动定位表头行：扫描含字段关键词的行
  2. 列名模糊匹配：同一含义多种叫法（成交日期/交易日期/日期...）
  3. 列映射可被前端修正后重新导入
  4. 配对：按证券代码分组 + 方向状态机，把连续成交配成完整交易（含加仓/减仓）
"""
import csv
import io
import re
from datetime import datetime

# ---------- 字段别名库（同一含义多种叫法） ----------
FIELD_ALIASES = {
    "datetime": ["成交日期", "交易日期", "发生日期", "日期", "业务日期", "委托日期"],
    "time": ["成交时间", "交易时间", "时间"],
    "code": ["证券代码", "证券编码", "股票代码", "合约代码", "标的代码", "代码", "合约", "证券号码"],
    "name": ["证券名称", "股票名称", "合约名称", "标的名称", "名称", "证券简称"],
    "direction": ["买卖标志", "操作", "买卖方向", "方向", "业务名称", "交易类型", "买卖类别",
                  "业务", "买/卖", "交易方向"],
    "price": ["成交价格", "价格", "成交价", "成交均价", "成交平均价"],
    "volume": ["成交数量", "数量", "成交量", "成交股数", "手数"],
    "amount": ["成交金额", "发生金额", "成交额", "金额", "资金发生额", "发生额"],
    "fee": ["手续费", "佣金", "费用", "交易费用", "结算费", "其他杂费", "其他费用"],
    "stamp_tax": ["印花税"],
    "transfer_fee": ["过户费", "过户"],
    "close_pnl": ["平仓盈亏", "平仓盈亏金额", "浮动盈亏"],
}

# 标准字段顺序（前端映射下拉用）
STANDARD_FIELDS = ["datetime", "time", "code", "name", "direction", "price", "volume",
                   "amount", "fee", "stamp_tax", "transfer_fee", "close_pnl"]

class ImporterError(RuntimeError):
    pass


# ---------- 文件读取 ----------
def _read_cells(path: str) -> list[list[str]]:
    """读取文件为二维字符串数组"""
    ext = path.rsplit(".", 1)[-1].lower() if "." in path else ""
    if ext in ("xlsx", "xlsm"):
        return _read_xlsx(path)
    if ext in ("xls",):
        raise ImporterError(
            "检测到旧版 Excel(.xls) 文件，请用 Excel 打开后另存为 .xlsx 或 .csv 再导入"
        )
    return _read_text(path)


def _read_xlsx(path: str) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImporterError("缺少 openpyxl 依赖，无法解析 Excel 文件")
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        raise ImporterError("Excel 文件解析失败，请确认是有效的 .xlsx 文件")
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = [_norm_cell(v) for v in row]
        rows.append(cells)
    wb.close()
    return rows


def _read_text(path: str) -> list[list[str]]:
    """读取 csv/txt，自动检测编码与分隔符"""
    raw = open(path, "rb").read()
    enc = _detect_encoding(raw)
    try:
        text = raw.decode(enc)
    except Exception:
        text = raw.decode("utf-8", errors="replace")
    # 去掉 BOM
    text = text.lstrip("\ufeff")
    lines = [l for l in text.splitlines() if l.strip()]
    if not lines:
        raise ImporterError("文件内容为空")
    # 检测分隔符：制表符 > 逗号 > 分号 > 空格
    sample = "\n".join(lines[:5])
    if "\t" in sample:
        delim = "\t"
    elif ";" in sample and "," not in sample:
        delim = ";"
    elif "," in sample:
        delim = ","
    elif re.search(r"\s{2,}", sample):
        delim = None  # 多空格分隔
    else:
        delim = ","
    rows = []
    for line in lines:
        if delim == ",":
            cells = next(csv.reader([line]))
        elif delim == "\t":
            cells = line.split("\t")
        elif delim == ";":
            cells = line.split(";")
        else:
            cells = [c for c in re.split(r"\s+", line.strip()) if c]
        rows.append([_norm_cell(c) for c in cells])
    return rows


def _detect_encoding(raw: bytes) -> str:
    if raw.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    # 先试 UTF-8（有严格字节校验，GBK 文件解码会失败）；GBK 几乎任何字节都能解码成功
    for enc in ("utf-8", "gbk"):
        try:
            raw.decode(enc)
            return enc
        except Exception:
            continue
    return "gbk"


def _norm_cell(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    s = str(v).strip()
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return s


# ---------- 表头定位与列映射 ----------
def _match_field(header_text: str):
    """根据表头文本匹配标准字段"""
    h = header_text.strip().lower().replace(" ", "").replace("_", "")
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if alias.lower() in h or h in alias.lower():
                return field
    return None


def locate_header(rows: list[list[str]], max_scan=20) -> int:
    """定位表头行（含最多字段关键词的行）"""
    best_idx, best_score = -1, 0
    for i, row in enumerate(rows[:max_scan]):
        score = sum(1 for c in row if _match_field(c))
        if score > best_score:
            best_score, best_idx = score, i
    if best_idx < 0 or best_score < 2:
        raise ImporterError(
            "未识别到交割单表头。请确认文件包含：成交日期/证券代码/买卖方向/成交价格等列"
        )
    return best_idx


def build_mapping(header_row: list[str]) -> dict:
    """表头行 → 列映射 {field: col_index}"""
    mapping = {}
    for col, cell in enumerate(header_row):
        field = _match_field(cell)
        if field and field not in mapping:
            mapping[field] = col
    return mapping


def extract_rows(rows: list[list[str]], header_idx: int) -> list[list[str]]:
    """提取表头之后的有效数据行"""
    data = []
    for row in rows[header_idx + 1:]:
        if any(str(c).strip() for c in row):
            data.append(row)
    return data


# ---------- 方向识别 ----------
# 包含匹配关键词（长词优先），覆盖同花顺"证券买入/证券卖出"、期货"买开/卖平"、英文等
BUY_KEYS = ("买入", "买开", "买平", "做多", "多开", "多平", "buy", "b")
SELL_KEYS = ("卖出", "卖开", "卖平", "做空", "空开", "空平", "sell", "s")


def parse_direction(text: str):
    """识别买卖方向，返回 'buy'/'sell'。支持各种叫法：证券买入/卖出、买开/卖平、B/S、Buy/Sell"""
    t = str(text or "").strip()
    if not t:
        return None
    tl = t.lower()
    for kw in BUY_KEYS:
        if kw in tl:
            return "buy"
    for kw in SELL_KEYS:
        if kw in tl:
            return "sell"
    # 兜底：含"买"→buy、含"卖"→sell（注意"卖"不含"买"）
    if "买" in tl:
        return "buy"
    if "卖" in tl:
        return "sell"
    if "多" in tl:
        return "buy"
    if "空" in tl:
        return "sell"
    return None


def _to_float(v):
    s = str(v or "").replace(",", "").replace("￥", "").replace("¥", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_positive_float(v):
    """只接受正数（成交价、成交数量），0/负数/无效→None"""
    n = _to_float(v)
    if n is None or n <= 0:
        return None
    return n


def _to_datetime(v):
    s = str(v or "").strip()
    if not s:
        return None
    # 统一各种分隔符
    s = s.replace("/", "-").replace(".", "-").replace("年", "-").replace("月", "-").replace("日", " ")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S.%f",
                "%Y-%m-%d", "%Y%m%d %H:%M:%S", "%Y%m%d"):
        try:
            return datetime.strptime(s.strip(), fmt)
        except ValueError:
            continue
    return None


# ---------- 解析主入口 ----------
def parse_file(path: str) -> dict:
    """解析交割单文件，返回：{headers, rows(前20条预览), mapping, total, missing_required}"""
    rows = _read_cells(path)
    header_idx = locate_header(rows)
    header = rows[header_idx]
    mapping = build_mapping(header)
    data = extract_rows(rows, header_idx)
    # 必要字段缺失时提示（允许前端进入手动列映射修正）
    required = {"datetime", "code", "direction", "price", "volume"}
    missing_required = [f for f in required if f not in mapping]
    return {
        "headers": header,
        "rows": data[:20],
        "mapping": mapping,
        "total": len(data),
        "missing_required": missing_required,
    }


def normalize_records(data: list[list[str]], mapping: dict) -> tuple[list[dict], dict]:
    """把数据行按映射转成标准记录 dict。

    返回 (records, skip_summary)：skip_summary 统计每类原因跳过的行数，便于诊断
    """
    records = []
    skip = {"rows": len(data), "no_direction": 0, "no_price": 0, "no_volume": 0,
            "no_datetime": 0, "no_code": 0}
    for row in data:
        def _g(field):
            col = mapping.get(field)
            return row[col] if col is not None and col < len(row) else ""
        direction = parse_direction(_g("direction"))
        # 价格/数量必须 > 0（撤单/无效行通常为 0 或空）
        price = _to_positive_float(_g("price"))
        volume = _to_positive_float(_g("volume"))
        code = _g("code").strip()
        # 日期+时间 合并
        date_s = _g("datetime").strip()
        time_s = _g("time").strip()
        dt = _to_datetime(f"{date_s} {time_s}" if time_s else date_s)
        if not direction:
            skip["no_direction"] += 1
            continue
        if price is None:
            skip["no_price"] += 1
            continue
        if volume is None:
            skip["no_volume"] += 1
            continue
        if dt is None:
            skip["no_datetime"] += 1
            continue
        if not code:
            skip["no_code"] += 1
            continue
        fee = sum(f for f in [_to_float(_g("fee")), _to_float(_g("stamp_tax")),
                              _to_float(_g("transfer_fee"))] if f is not None)
        records.append({
            "code": code,
            "name": _g("name").strip(),
            "datetime": dt,
            "direction": direction,  # buy/sell
            "price": price,
            "volume": volume,
            "amount": _to_float(_g("amount")),
            "fee": fee,
            "close_pnl": _to_float(_g("close_pnl")),
        })
    records.sort(key=lambda r: (r["code"], r["datetime"]))
    return records, skip


# ---------- 配对算法：把成交明细配成完整交易 ----------
def pair_trades(records: list[dict]) -> list[dict]:
    """
    按证券代码分组，组内按时间顺序状态机配对：
      - 无持仓时首笔开仓（买→long，卖→short）
      - 同向成交 → 加仓（position_actions: add）
      - 反向成交 → 减仓/平仓（position_actions: reduce / 最后一笔平仓）
      - 反向成交数量超过持仓 → 平仓并反向开新仓
    返回完整交易列表
    """
    from collections import OrderedDict

    groups = OrderedDict()
    for r in records:
        groups.setdefault(r["code"], []).append(r)

    trades = []
    for code, fills in groups.items():
        current = None  # 当前在建交易
        for f in fills:
            is_buy = f["direction"] == "buy"
            if current is None:
                current = _open_trade(f)
                continue
            same_dir = (is_buy and current["direction"] == "long") or \
                       (not is_buy and current["direction"] == "short")
            if same_dir:
                current = _add_fill(current, f)
                continue
            # 反向：平仓或开反向
            remaining = current["_remaining"]
            if f["volume"] <= remaining:
                current = _reduce_fill(current, f)
                if current["_remaining"] <= 0:
                    _finalize(current)
                    trades.append(current)
                    current = None
            else:
                # 平掉剩余 + 剩余数量反向开新仓
                over = f["volume"] - remaining
                part1 = {**f, "volume": remaining}
                current = _reduce_fill(current, part1)
                _finalize(current)
                trades.append(current)
                part2 = {**f, "volume": over}
                current = _open_trade(part2)
        # 未平仓的持仓也作为一笔（exit=最后一次成交，标记未平）
        if current is not None and current["_remaining"] > 0:
            _finalize(current)
            current["unclosed"] = True
            trades.append(current)
    return trades


def _open_trade(f: dict) -> dict:
    is_buy = f["direction"] == "buy"
    amount = abs(f.get("amount") or 0.0)
    return {
        "code": f["code"],
        "name": f["name"],
        "direction": "long" if is_buy else "short",
        "entry_time": f["datetime"],
        "entry_price": f["price"],
        "exit_time": f["datetime"],
        "exit_price": f["price"],
        "volume": f["volume"],
        "fee": f["fee"],
        "pnl_raw": f.get("close_pnl") or 0.0,
        "cost": amount if is_buy else 0.0,   # 买入成本
        "revenue": amount if not is_buy else 0.0,  # 卖出收入
        "actions": [],  # [{action_type, time, price, volume}]
        "_remaining": f["volume"],
    }


def _add_fill(t: dict, f: dict) -> dict:
    """加仓：数量增加，更新最新价格/时间"""
    t["volume"] += f["volume"]
    t["_remaining"] += f["volume"]
    t["exit_time"] = f["datetime"]
    t["exit_price"] = f["price"]
    t["fee"] += f["fee"]
    t["pnl_raw"] += (f.get("close_pnl") or 0.0)
    amount = abs(f.get("amount") or 0.0)
    if f["direction"] == "buy":
        t["cost"] += amount
    else:
        t["revenue"] += amount
    t["actions"].append({
        "action_type": "add",
        "action_time": f["datetime"],
        "price": f["price"],
        "volume": f["volume"],
    })
    return t


def _reduce_fill(t: dict, f: dict) -> dict:
    t["_remaining"] -= f["volume"]
    t["exit_time"] = f["datetime"]
    t["exit_price"] = f["price"]
    t["fee"] += f["fee"]
    t["pnl_raw"] += (f.get("close_pnl") or 0.0)
    amount = abs(f.get("amount") or 0.0)
    if f["direction"] == "buy":
        t["cost"] += amount
    else:
        t["revenue"] += amount
    t["actions"].append({
        "action_type": "reduce",
        "action_time": f["datetime"],
        "price": f["price"],
        "volume": f["volume"],
    })
    return t


def _finalize(t: dict):
    """计算最终 pnl（不含手续费，手续费单独存 fee 字段展示）"""
    # 期货交割单用平仓盈亏；否则 卖出收入-买入成本
    if t["pnl_raw"]:
        t["pnl"] = round(t["pnl_raw"], 2)
    elif t["cost"] or t["revenue"]:
        t["pnl"] = round(t["revenue"] - t["cost"], 2)
    else:
        t["pnl"] = None
    t.pop("_remaining", None)


# ---------- 重复检测 ----------
def duplicate_key(trade: dict) -> tuple:
    """生成交易唯一键，用于去重（代码+入场时间+入场价+方向）"""
    return (
        trade["code"],
        trade["entry_time"].strftime("%Y-%m-%d %H:%M:%S"),
        round(trade["entry_price"], 4),
        trade["direction"],
    )


def fill_fingerprint(rec: dict) -> str:
    """成交记录指纹（去重用）"""
    return f"{rec['code']}|{rec['datetime'].strftime('%Y-%m-%d %H:%M:%S')}|{rec['price']}|{rec['volume']}|{rec['direction']}"


def infer_instrument_type(code: str) -> str:
    """根据证券代码推断品种类型：数字→A股，字母+数字→商品期货，纯字母→数字货币"""
    if re.fullmatch(r"\d+", code or ""):
        return "A股"
    if re.fullmatch(r"[A-Za-z0-9]+", code or "") and re.search(r"[A-Za-z]", code or ""):
        return "商品期货"
    return "数字货币"


def merge_incremental(db, records: list[dict], user_id: int) -> dict:
    """
    增量合并：把成交明细按时间顺序与数据库中未平仓交易逐条合并。

    规则：
      - 该品种无未平仓交易 → 新建交易（记录开仓时间/价格/数量）
      - 有未平仓交易：
          * 同方向成交 → 加仓（累加数量，记加仓动作）
          * 反方向成交 → 平仓/部分平仓（记减仓动作；完全平仓后计算盈亏）
          * 反向成交量超出持仓 → 平掉持仓 + 剩余反向开新仓
      - 已处理过的成交（指纹重复）→ 跳过，防止重复导入重复合并

    返回 {imported_new, merged_actions, skipped}
    """
    from ..models import ImportFill, Trade, TradePositionAction

    imported_new = 0
    merged_actions = 0
    skipped = 0

    # 按 code 分组（records 已按 code+时间排序，这里用分组保证同 code 连续处理）
    groups = {}
    for r in records:
        groups.setdefault(r["code"], []).append(r)

    for code, fills in groups.items():
        for rec in fills:
            fp = fill_fingerprint(rec)
            if db.query(ImportFill.id).filter_by(user_id=user_id, fingerprint=fp).first():
                skipped += 1
                continue

            # 找该品种当前未平仓交易
            open_trade = (
                db.query(Trade)
                .filter(Trade.user_id == user_id, Trade.instrument_code == code,
                        Trade.remaining_volume > 0)
                .order_by(Trade.entry_time.asc())
                .first()
            )
            is_buy = rec["direction"] == "buy"

            if open_trade is None:
                # 场景1/2：新开仓（出场留空，待完全平仓时记录）
                trade = Trade(
                    user_id=user_id,
                    instrument_type=infer_instrument_type(code),
                    instrument_code=code,
                    instrument_name=rec.get("name") or code,
                    direction="long" if is_buy else "short",
                    entry_time=rec["datetime"],
                    entry_price=rec["price"],
                    exit_time=None,
                    exit_price=None,
                    volume=rec["volume"],
                    remaining_volume=rec["volume"],
                    fee=rec["fee"],
                    import_cost=(rec.get("amount") or 0.0) if is_buy else 0.0,
                    import_revenue=0.0 if is_buy else (rec.get("amount") or 0.0),
                    pnl=None,
                )
                db.add(trade)
                db.flush()
                imported_new += 1
            else:
                same_dir = (is_buy and open_trade.direction == "long") or \
                           (not is_buy and open_trade.direction == "short")
                amount = abs(rec.get("amount") or 0.0)
                if same_dir:
                    # 场景3/4：加仓（不更新出场，出场只在完全平仓时记录）
                    open_trade.volume += rec["volume"]
                    open_trade.remaining_volume += rec["volume"]
                    open_trade.fee += rec["fee"]
                    if is_buy:
                        open_trade.import_cost += amount
                    else:
                        open_trade.import_revenue += amount
                    _add_position_action(db, open_trade, rec, "add")
                    merged_actions += 1
                else:
                    remaining = open_trade.remaining_volume
                    if rec["volume"] <= remaining:
                        # 场景5/6：部分平仓（出场留空） / 全部平仓（记录出场）
                        open_trade.remaining_volume = remaining - rec["volume"]
                        open_trade.fee += rec["fee"]
                        if is_buy:
                            open_trade.import_cost += amount
                        else:
                            open_trade.import_revenue += amount
                        _add_position_action(db, open_trade, rec, "reduce")
                        merged_actions += 1
                        if open_trade.remaining_volume <= 0:
                            open_trade.exit_time = rec["datetime"]
                            open_trade.exit_price = rec["price"]
                            _finalize_pnl(open_trade)
                    else:
                        # 场景7：平掉全部持仓 + 剩余反向开新仓
                        over = rec["volume"] - remaining
                        part = {**rec, "volume": remaining}
                        open_trade.remaining_volume = 0
                        open_trade.exit_time = part["datetime"]
                        open_trade.exit_price = part["price"]
                        open_trade.fee += part["fee"]
                        if part["direction"] == "buy":
                            open_trade.import_cost += abs(part.get("amount") or 0.0)
                        else:
                            open_trade.import_revenue += abs(part.get("amount") or 0.0)
                        _add_position_action(db, open_trade, part, "reduce")
                        _finalize_pnl(open_trade)
                        # 剩余反向开新仓（出场留空）
                        trade = Trade(
                            user_id=user_id,
                            instrument_type=infer_instrument_type(code),
                            instrument_code=code,
                            instrument_name=rec.get("name") or code,
                            direction="short" if open_trade.direction == "long" else "long",
                            entry_time=rec["datetime"],
                            entry_price=rec["price"],
                            exit_time=None,
                            exit_price=None,
                            volume=over,
                            remaining_volume=over,
                            fee=rec["fee"],
                            import_cost=0.0,
                            import_revenue=0.0,
                            pnl=None,
                        )
                        if not is_buy:
                            trade.import_cost = 0.0
                            trade.import_revenue = amount if rec["direction"] == "sell" else 0.0
                        db.add(trade)
                        db.flush()
                        imported_new += 1

            # 记录指纹
            db.add(ImportFill(user_id=user_id, fingerprint=fp))

    return {"imported_new": imported_new, "merged_actions": merged_actions, "skipped": skipped}


def _add_position_action(db, trade, rec, action_type: str):
    """记录加仓/减仓动作（volume 正=加仓 负=减仓）"""
    from ..models import TradePositionAction
    count = db.query(TradePositionAction).filter(TradePositionAction.trade_id == trade.id).count()
    signed = rec["volume"] if action_type == "add" else -rec["volume"]
    db.add(TradePositionAction(
        trade_id=trade.id,
        action_time=rec["datetime"],
        price=rec["price"],
        volume=signed,
        note="",
        sort_order=count,
    ))


def _finalize_pnl(trade):
    """完全平仓后计算盈亏（不含手续费）"""
    if trade.import_cost or trade.import_revenue:
        trade.pnl = round(trade.import_revenue - trade.import_cost, 2)
    else:
        trade.pnl = None
